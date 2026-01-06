import os
import vtk
import qt
import ctk
import slicer
from slicer.ScriptedLoadableModule import *
from slicer.util import VTKObservationMixin
import numpy as np

class VolBrainVolumeCalculator(ScriptedLoadableModule):
    """Base class for 3D Slicer module."""
    
    def __init__(self, parent):
        ScriptedLoadableModule.__init__(self, parent)
        self.parent.title = "VolBrain Volume Calculator"
        self.parent.categories = ["Quantification"]
        self.parent.dependencies = []
        self.parent.contributors = ["Niyazi Acer (Sanko University)"]
        self.parent.helpText = """
3D Slicer extension for comprehensive brain structure volume calculation 
and 3D visualization from volBrain segmentation outputs. Supports 108+ 
cortical regions, tissue classification, lobes, and macrostructures.
"""
        self.parent.acknowledgementText = """volBrain-based volume analysis module"""

class VolBrainVolumeCalculatorWidget(ScriptedLoadableModuleWidget, VTKObservationMixin):
    """Module user interface."""
    
    def __init__(self, parent=None):
        ScriptedLoadableModuleWidget.__init__(self, parent)
        VTKObservationMixin.__init__(self)
        self.logic = None
        self.volumeResults = {}
        self.loadedNodes = []
        self.currentSegmentationNode = None
        self.currentOpacity = 1.0
        
    def setup(self):
        """Creates interface components."""
        ScriptedLoadableModuleWidget.setup(self)
        
        self.logic = VolBrainVolumeCalculatorLogic()
        
        # Input Files
        inputsCollapsibleButton = ctk.ctkCollapsibleButton()
        inputsCollapsibleButton.text = "Input Files"
        self.layout.addWidget(inputsCollapsibleButton)
        inputsFormLayout = qt.QFormLayout(inputsCollapsibleButton)
        
        self.structuresSelector = ctk.ctkPathLineEdit()
        self.structuresSelector.nameFilters = ["NIfTI (*.nii *.nii.gz)"]
        inputsFormLayout.addRow("Structures:", self.structuresSelector)

        self.tissuesSelector = ctk.ctkPathLineEdit()
        self.tissuesSelector.nameFilters = ["NIfTI (*.nii *.nii.gz)"]
        inputsFormLayout.addRow("Tissues:", self.tissuesSelector)

        self.lobesSelector = ctk.ctkPathLineEdit()
        self.lobesSelector.nameFilters = ["NIfTI (*.nii *.nii.gz)"]
        inputsFormLayout.addRow("Lobes:", self.lobesSelector)

        self.macroSelector = ctk.ctkPathLineEdit()
        self.macroSelector.nameFilters = ["NIfTI (*.nii *.nii.gz)"]
        inputsFormLayout.addRow("Macrostructures:", self.macroSelector)
             
        self.quickLoadButton = qt.QPushButton("Load Automatically from Folder")
        inputsFormLayout.addRow(self.quickLoadButton)
        
        # Calculation
        calcCollapsibleButton = ctk.ctkCollapsibleButton()
        calcCollapsibleButton.text = "Calculation and Visualization"
        self.layout.addWidget(calcCollapsibleButton)
        calcFormLayout = qt.QFormLayout(calcCollapsibleButton)
        
        self.progressBar = qt.QProgressBar()
        self.progressBar.setRange(0, 100)
        self.progressBar.setValue(0)
        calcFormLayout.addRow("Progress:", self.progressBar)
        
        self.statusLabel = qt.QLabel("Ready")
        calcFormLayout.addRow("Status:", self.statusLabel)
        
        # 3D visualization options
        self.show3DCheckbox = qt.QCheckBox("3D Visualization")
        self.show3DCheckbox.checked = True
        calcFormLayout.addRow(self.show3DCheckbox)
        
        self.applyButton = qt.QPushButton("Calculate Volumes and Visualize")
        self.applyButton.setStyleSheet("QPushButton { font-weight: bold; padding: 10px; }")
        calcFormLayout.addRow(self.applyButton)
        
        # Results
        resultsCollapsibleButton = ctk.ctkCollapsibleButton()
        resultsCollapsibleButton.text = "Results"
        resultsCollapsibleButton.collapsed = False
        self.layout.addWidget(resultsCollapsibleButton)
        resultsFormLayout = qt.QFormLayout(resultsCollapsibleButton)
        
        self.resultsTable = qt.QTableWidget()
        self.resultsTable.setColumnCount(5)
        self.resultsTable.setHorizontalHeaderLabels(["Category", "Label ID", "Structure Name", "Volume (mm³)", "Volume (ml)"])
        self.resultsTable.setColumnWidth(0, 120)
        self.resultsTable.setColumnWidth(1, 80)
        self.resultsTable.setColumnWidth(2, 250)
        self.resultsTable.setColumnWidth(3, 120)
        self.resultsTable.setColumnWidth(4, 100)
        self.resultsTable.setAlternatingRowColors(True)
        self.resultsTable.setSortingEnabled(True)
        self.resultsTable.setMinimumHeight(400)
        resultsFormLayout.addRow(self.resultsTable)
        
        self.summaryLabel = qt.QLabel("")
        self.summaryLabel.setWordWrap(True)
        resultsFormLayout.addRow("Summary:", self.summaryLabel)
        
        # Export buttons
        exportLayout = qt.QHBoxLayout()
        self.exportCSVButton = qt.QPushButton("💾 CSV")
        self.exportCSVButton.enabled = False
        self.exportCSVButton.setMaximumWidth(100)
        
        self.exportExcelButton = qt.QPushButton("📊 Excel")
        self.exportExcelButton.enabled = False
        self.exportExcelButton.setMaximumWidth(100)
        
        self.copyButton = qt.QPushButton("📋 Copy")
        self.copyButton.enabled = False
        self.copyButton.setMaximumWidth(100)
        
        self.clearButton = qt.QPushButton("🗑️ Clear")
        self.clearButton.enabled = False
        self.clearButton.setMaximumWidth(100)
        
        exportLayout.addWidget(self.exportCSVButton)
        exportLayout.addWidget(self.exportExcelButton)
        exportLayout.addWidget(self.copyButton)
        exportLayout.addWidget(self.clearButton)
        exportLayout.addStretch()
        resultsFormLayout.addRow(exportLayout)
        
        # 3D Visualization Controls
        visualCollapsibleButton = ctk.ctkCollapsibleButton()
        visualCollapsibleButton.text = "3D Visualization Controls"
        visualCollapsibleButton.collapsed = True
        self.layout.addWidget(visualCollapsibleButton)
        visualFormLayout = qt.QFormLayout(visualCollapsibleButton)
        
        # Segment selector
        self.segmentSelector = qt.QComboBox()
        self.segmentSelector.addItem("-- All Structures --")
        self.segmentSelector.enabled = False
        visualFormLayout.addRow("Show in 3D:", self.segmentSelector)
        
        # Visibility buttons
        visibilityLayout = qt.QHBoxLayout()
        self.showAllButton = qt.QPushButton("👁️ Show All")
        self.showAllButton.enabled = False
        self.hideAllButton = qt.QPushButton("🚫 Hide All")
        self.hideAllButton.enabled = False
        self.toggleOpacityButton = qt.QPushButton("🌓 Toggle Opacity")
        self.toggleOpacityButton.enabled = False
        visibilityLayout.addWidget(self.showAllButton)
        visibilityLayout.addWidget(self.hideAllButton)
        visibilityLayout.addWidget(self.toggleOpacityButton)
        visualFormLayout.addRow(visibilityLayout)
        
        # Connections
        self.quickLoadButton.connect('clicked(bool)', self.onQuickLoad)
        self.applyButton.connect('clicked(bool)', self.onApplyButton)
        self.exportCSVButton.connect('clicked(bool)', self.onExportCSV)
        self.exportExcelButton.connect('clicked(bool)', self.onExportExcel)
        self.copyButton.connect('clicked(bool)', self.onCopyToClipboard)
        self.clearButton.connect('clicked(bool)', self.onClear)
        self.segmentSelector.connect('currentIndexChanged(int)', self.onSegmentSelected)
        self.showAllButton.connect('clicked(bool)', self.onShowAll)
        self.hideAllButton.connect('clicked(bool)', self.onHideAll)
        self.toggleOpacityButton.connect('clicked(bool)', self.onToggleOpacity)
        
        self.layout.addStretch(1)
        
    def cleanup(self):
        """Cleanup operations."""
        pass
    
    def onQuickLoad(self):
        """Automatic file loading from folder."""
        folder = qt.QFileDialog.getExistingDirectory(self.parent, "Select volBrain Output Folder")
        
        if folder:
            files = os.listdir(folder)
            for f in files:
                fullPath = os.path.join(folder, f)
                if 'structures' in f.lower() and 'native' in f.lower() and not 'macro' in f.lower():
                    self.structuresSelector.setCurrentPath(fullPath)
                elif 'tissues' in f.lower() and 'native' in f.lower():
                    self.tissuesSelector.setCurrentPath(fullPath)
                elif 'lobes' in f.lower() and 'native' in f.lower():
                    self.lobesSelector.setCurrentPath(fullPath)
                elif 'macrostructures' in f.lower() and 'native' in f.lower():
                    self.macroSelector.setCurrentPath(fullPath)
            
            self.statusLabel.setText("Files loaded automatically")
    
    def onApplyButton(self):
        """Starts volume calculation process."""
        self.volumeResults = {}
        self.resultsTable.setRowCount(0)
        self.progressBar.setValue(0)
        self.statusLabel.setText("Starting calculation...")
        slicer.app.processEvents()
        
        # Collect file paths
        files = [
            (self.structuresSelector.currentPath, "structures"),
            (self.tissuesSelector.currentPath, "tissues"),
            (self.lobesSelector.currentPath, "lobes"),
            (self.macroSelector.currentPath, "macro")
        ]
        
        validFiles = [(path, cat) for path, cat in files if os.path.exists(path)]
        
        if not validFiles:
            slicer.util.errorDisplay("Please select at least one file!")
            self.statusLabel.setText("Error: No file selected")
            return
        
        totalSteps = len(validFiles)
        currentStep = 0
        allResults = {}
        
        for filePath, category in validFiles:
            currentStep += 1
            self.progressBar.setValue(int(currentStep / totalSteps * 100))
            self.statusLabel.setText(f"Calculating {category}...")
            slicer.app.processEvents()
            
            try:
                results = self.logic.calculateVolumes(
                    filePath, 
                    category, 
                    show3D=self.show3DCheckbox.checked
                )
                allResults.update(results)
                
                # Save node
                if self.show3DCheckbox.checked:
                    # Find and save segmentation node
                    segNodeName = f"volBrain_{category}_Segmentation"
                    segNode = slicer.util.getFirstNodeByName(segNodeName)
                    if segNode:
                        self.loadedNodes.append(segNode)
                        if not self.currentSegmentationNode:
                            self.currentSegmentationNode = segNode
                
            except Exception as e:
                print(f"ERROR {category}: {str(e)}")
                import traceback
                traceback.print_exc()
        
        self.volumeResults = allResults
        self.updateResultsTable()
        self.updateSummary()
        
        self.exportCSVButton.enabled = True
        self.exportExcelButton.enabled = True
        self.copyButton.enabled = True
        self.clearButton.enabled = True
        self.segmentSelector.enabled = True
        self.showAllButton.enabled = True
        self.hideAllButton.enabled = True
        self.toggleOpacityButton.enabled = True
        
        # Fill segment selector
        self.updateSegmentSelector()
        
        self.progressBar.setValue(100)
        self.statusLabel.setText(f"Complete! {len(allResults)} structures calculated")
        
        # Setup 3D view
        if self.show3DCheckbox.checked:
            layoutManager = slicer.app.layoutManager()
            layoutManager.setLayout(slicer.vtkMRMLLayoutNode.SlicerLayoutFourUpView)
            slicer.util.resetSliceViews()
        
        slicer.util.messageBox(f"Volume calculation complete!\n\n{len(allResults)} brain structures analyzed.")
    
    def updateResultsTable(self):
        """Updates results table."""
        self.resultsTable.setRowCount(len(self.volumeResults))
        row = 0
        
        for structureKey, data in sorted(self.volumeResults.items()):
            self.resultsTable.setItem(row, 0, qt.QTableWidgetItem(data['category'].upper()))
            self.resultsTable.setItem(row, 1, qt.QTableWidgetItem(str(data['label_id'])))
            self.resultsTable.setItem(row, 2, qt.QTableWidgetItem(data['name']))
            self.resultsTable.setItem(row, 3, qt.QTableWidgetItem(f"{data['mm3']:.2f}"))
            self.resultsTable.setItem(row, 4, qt.QTableWidgetItem(f"{data['ml']:.4f}"))
            row += 1
    
    def updateSummary(self):
        """Updates summary statistics."""
        if not self.volumeResults:
            return
        
        categories = {}
        for data in self.volumeResults.values():
            cat = data['category']
            if cat not in categories:
                categories[cat] = {'count': 0, 'total_ml': 0}
            categories[cat]['count'] += 1
            categories[cat]['total_ml'] += data['ml']
        
        summary = "<b>Category Summary:</b><br>"
        for cat, stats in sorted(categories.items()):
            summary += f"{cat.upper()}: {stats['count']} structures, Total: {stats['total_ml']:.2f} ml<br>"
        
        self.summaryLabel.setText(summary)
    
    def onExportCSV(self):
        """Export in CSV format."""
        fileName = qt.QFileDialog.getSaveFileName(
            self.parent, "Save CSV File", 
            os.path.expanduser("~/volbrain_volumes.csv"), 
            "CSV Files (*.csv)")
        
        if fileName:
            try:
                with open(fileName, 'w', encoding='utf-8') as f:
                    f.write("Category,Label_ID,Structure_Name,Volume_mm3,Volume_ml\n")
                    for data in sorted(self.volumeResults.values(), key=lambda x: (x['category'], x['label_id'])):
                        f.write(f"{data['category']},{data['label_id']},{data['name']},{data['mm3']:.2f},{data['ml']:.4f}\n")
                
                slicer.util.messageBox(f"Results saved:\n{fileName}")
                self.statusLabel.setText("CSV saved")
            except Exception as e:
                slicer.util.errorDisplay(f"Save error: {str(e)}")
    
    def onCopyToClipboard(self):
        """Copy results to clipboard."""
        text = "Category\tLabel_ID\tStructure_Name\tVolume_mm3\tVolume_ml\n"
        for data in sorted(self.volumeResults.values(), key=lambda x: (x['category'], x['label_id'])):
            text += f"{data['category']}\t{data['label_id']}\t{data['name']}\t{data['mm3']:.2f}\t{data['ml']:.4f}\n"
        
        clipboard = qt.QApplication.clipboard()
        clipboard.setText(text)
        self.statusLabel.setText("Copied to clipboard")
        slicer.util.messageBox("Results copied to clipboard!")
    
    def onClear(self):
        """Clear loaded nodes."""
        for node in self.loadedNodes:
            slicer.mrmlScene.RemoveNode(node)
        self.loadedNodes = []
        self.currentSegmentationNode = None
        self.segmentSelector.clear()
        self.segmentSelector.addItem("-- All Structures --")
        self.segmentSelector.enabled = False
        self.showAllButton.enabled = False
        self.hideAllButton.enabled = False
        self.toggleOpacityButton.enabled = False
        self.statusLabel.setText("Cleared")
    
    def updateSegmentSelector(self):
        """Update segment selector."""
        self.segmentSelector.clear()
        self.segmentSelector.addItem("-- All Structures --")
        
        # Find all segmentation nodes
        for node in self.loadedNodes:
            if node.IsA('vtkMRMLSegmentationNode'):
                self.currentSegmentationNode = node
                segmentation = node.GetSegmentation()
                for i in range(segmentation.GetNumberOfSegments()):
                    segmentId = segmentation.GetNthSegmentID(i)
                    segment = segmentation.GetSegment(segmentId)
                    self.segmentSelector.addItem(segment.GetName(), segmentId)
    
    def onSegmentSelected(self, index):
        """Show selected segment in 3D."""
        if not self.currentSegmentationNode:
            return
        
        segmentation = self.currentSegmentationNode.GetSegmentation()
        displayNode = self.currentSegmentationNode.GetDisplayNode()
        
        if index == 0:  # "All Structures" selected
            # Show all
            for i in range(segmentation.GetNumberOfSegments()):
                segmentId = segmentation.GetNthSegmentID(i)
                displayNode.SetSegmentVisibility3D(segmentId, True)
        else:
            # Show only selected
            selectedSegmentId = self.segmentSelector.itemData(index)
            for i in range(segmentation.GetNumberOfSegments()):
                segmentId = segmentation.GetNthSegmentID(i)
                if segmentId == selectedSegmentId:
                    displayNode.SetSegmentVisibility3D(segmentId, True)
                else:
                    displayNode.SetSegmentVisibility3D(segmentId, False)
    
    def onShowAll(self):
        """Show all segments."""
        if not self.currentSegmentationNode:
            return
        
        segmentation = self.currentSegmentationNode.GetSegmentation()
        displayNode = self.currentSegmentationNode.GetDisplayNode()
        
        for i in range(segmentation.GetNumberOfSegments()):
            segmentId = segmentation.GetNthSegmentID(i)
            displayNode.SetSegmentVisibility3D(segmentId, True)
        
        self.segmentSelector.setCurrentIndex(0)
        self.statusLabel.setText("All structures visible")
    
    def onHideAll(self):
        """Hide all segments."""
        if not self.currentSegmentationNode:
            return
        
        segmentation = self.currentSegmentationNode.GetSegmentation()
        displayNode = self.currentSegmentationNode.GetDisplayNode()
        
        for i in range(segmentation.GetNumberOfSegments()):
            segmentId = segmentation.GetNthSegmentID(i)
            displayNode.SetSegmentVisibility3D(segmentId, False)
        
        self.statusLabel.setText("All structures hidden")
    
    def onToggleOpacity(self):
        """Toggle 3D view opacity."""
        if not self.currentSegmentationNode:
            return
        
        displayNode = self.currentSegmentationNode.GetDisplayNode()
        
        # Change opacity: 1.0 -> 0.5 -> 0.2 -> 1.0
        if self.currentOpacity >= 1.0:
            self.currentOpacity = 0.5
        elif self.currentOpacity >= 0.5:
            self.currentOpacity = 0.2
        else:
            self.currentOpacity = 1.0
        
        displayNode.SetOpacity3D(self.currentOpacity)
        self.statusLabel.setText(f"Opacity: {int(self.currentOpacity*100)}%")
    
    def onExportExcel(self):
        """Export in Excel format."""
        fileName = qt.QFileDialog.getSaveFileName(
            self.parent, "Save Excel File", 
            os.path.expanduser("~/volbrain_volumes.xlsx"), 
            "Excel Files (*.xlsx)")
        
        if fileName:
            try:
                # Check if openpyxl is available
                try:
                    import openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment
                    
                    # Create workbook
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "VolBrain Results"
                    
                    # Header styling
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    # Write headers
                    headers = ["Category", "Label ID", "Structure Name", "Volume (mm³)", "Volume (ml)"]
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Write data
                    row_num = 2
                    for data in sorted(self.volumeResults.values(), key=lambda x: (x['category'], x['label_id'])):
                        ws.cell(row=row_num, column=1, value=data['category'])
                        ws.cell(row=row_num, column=2, value=data['label_id'])
                        ws.cell(row=row_num, column=3, value=data['name'])
                        ws.cell(row=row_num, column=4, value=round(data['mm3'], 2))
                        ws.cell(row=row_num, column=5, value=round(data['ml'], 4))
                        row_num += 1
                    
                    # Adjust column widths
                    ws.column_dimensions['A'].width = 15
                    ws.column_dimensions['B'].width = 10
                    ws.column_dimensions['C'].width = 35
                    ws.column_dimensions['D'].width = 15
                    ws.column_dimensions['E'].width = 15
                    
                    # Add summary sheet
                    ws_summary = wb.create_sheet("Summary")
                    ws_summary.cell(row=1, column=1, value="Category").font = header_font
                    ws_summary.cell(row=1, column=2, value="Structure Count").font = header_font
                    ws_summary.cell(row=1, column=3, value="Total Volume (ml)").font = header_font
                    
                    categories = {}
                    for data in self.volumeResults.values():
                        cat = data['category']
                        if cat not in categories:
                            categories[cat] = {'count': 0, 'total_ml': 0}
                        categories[cat]['count'] += 1
                        categories[cat]['total_ml'] += data['ml']
                    
                    row_num = 2
                    for cat, stats in sorted(categories.items()):
                        ws_summary.cell(row=row_num, column=1, value=cat.upper())
                        ws_summary.cell(row=row_num, column=2, value=stats['count'])
                        ws_summary.cell(row=row_num, column=3, value=round(stats['total_ml'], 2))
                        row_num += 1
                    
                    ws_summary.column_dimensions['A'].width = 15
                    ws_summary.column_dimensions['B'].width = 18
                    ws_summary.column_dimensions['C'].width = 20
                    
                    # Save workbook
                    wb.save(fileName)
                    
                    slicer.util.messageBox(f"Excel file saved successfully:\n{fileName}")
                    self.statusLabel.setText("Excel file saved")
                    
                except ImportError:
                    # Fallback to CSV if openpyxl not available
                    slicer.util.warningDisplay(
                        "openpyxl module not found. Saving as CSV instead.\n\n"
                        "To enable Excel export, install openpyxl:\n"
                        "pip install openpyxl"
                    )
                    # Save as CSV with Excel-compatible format
                    fileName = fileName.replace('.xlsx', '.csv')
                    with open(fileName, 'w', encoding='utf-8-sig') as f:
                        f.write("Category\tLabel_ID\tStructure_Name\tVolume_mm3\tVolume_ml\n")
                        for data in sorted(self.volumeResults.values(), key=lambda x: (x['category'], x['label_id'])):
                            f.write(f"{data['category']}\t{data['label_id']}\t{data['name']}\t{data['mm3']:.2f}\t{data['ml']:.4f}\n")
                    
                    slicer.util.messageBox(f"Saved as CSV (Excel-compatible):\n{fileName}")
                    self.statusLabel.setText("CSV file saved")
                    
            except Exception as e:
                slicer.util.errorDisplay(f"Save error: {str(e)}\n\nDetails: {repr(e)}")
                import traceback
                traceback.print_exc()

class VolBrainVolumeCalculatorLogic(ScriptedLoadableModuleLogic):
    """Volume calculation logic."""
    
    def __init__(self):
        ScriptedLoadableModuleLogic.__init__(self)
    
    def calculateVolumes(self, filePath, category, show3D=True):
        """Calculate volumes from specified file."""
        
        nodeName = f"volBrain_{category}"
        
        # Remove old node if exists
        try:
            oldNode = slicer.util.getNode(nodeName)
            if oldNode:
                slicer.mrmlScene.RemoveNode(oldNode)
        except:
            pass  # No problem if node doesn't exist
        
        # Load NIfTI file
        volumeNode = slicer.util.loadVolume(filePath, returnNode=True)[1]
        volumeNode.SetName(nodeName)
        
        # Get voxel dimensions
        spacing = volumeNode.GetSpacing()
        voxelVolume = spacing[0] * spacing[1] * spacing[2]
        
        # Convert to label map
        labelNode = slicer.mrmlScene.AddNewNodeByClass('vtkMRMLLabelMapVolumeNode')
        labelNode.SetName(f"{nodeName}_labels")
        slicer.modules.volumes.logic().CreateLabelVolumeFromVolume(slicer.mrmlScene, labelNode, volumeNode)
        
        # Get array
        array = slicer.util.arrayFromVolume(labelNode)
        uniqueLabels = np.unique(array)
        uniqueLabels = uniqueLabels[uniqueLabels > 0]
        
        results = {}
        labelNames = self.getLabelNames(category)
        colorTable = self.getColorTable(category)
        
        # Create color table
        colorNode = slicer.mrmlScene.CreateNodeByClass('vtkMRMLColorTableNode')
        colorNode.SetName(f"{nodeName}_ColorTable")
        colorNode.SetTypeToUser()
        colorNode.SetNumberOfColors(int(np.max(uniqueLabels)) + 1)
        colorNode.NamesInitialisedOn()
        
        # Background
        colorNode.SetColor(0, "Background", 0.0, 0.0, 0.0, 0.0)
        
        for label in uniqueLabels:
            labelInt = int(label)
            voxelCount = np.sum(array == label)
            volumeMm3 = float(voxelCount * voxelVolume)
            volumeMl = volumeMm3 / 1000.0
            
            labelName = labelNames.get(labelInt, f"Label_{labelInt}")
            structureKey = f"{category}_{labelInt}_{labelName}"
            
            results[structureKey] = {
                "category": category,
                "label_id": labelInt,
                "name": labelName,
                "mm3": volumeMm3, 
                "ml": volumeMl
            }
            
            # Assign color
            if labelInt in colorTable:
                r, g, b = colorTable[labelInt]
                colorNode.SetColor(labelInt, labelName, r, g, b, 1.0)
            else:
                # Random color
                import random
                r, g, b = random.random(), random.random(), random.random()
                colorNode.SetColor(labelInt, labelName, r, g, b, 1.0)
        
        slicer.mrmlScene.AddNode(colorNode)
        
        # Assign color table to label node
        displayNode = labelNode.GetDisplayNode()
        if not displayNode:
            displayNode = slicer.mrmlScene.CreateNodeByClass('vtkMRMLLabelMapVolumeDisplayNode')
            slicer.mrmlScene.AddNode(displayNode)
            labelNode.SetAndObserveDisplayNodeID(displayNode.GetID())
        
        displayNode.SetAndObserveColorNodeID(colorNode.GetID())
        
        # 3D visualization
        if show3D:
            # Create segmentation
            segmentationNode = slicer.mrmlScene.AddNewNodeByClass('vtkMRMLSegmentationNode')
            segmentationNode.SetName(f"{nodeName}_Segmentation")
            
            # Convert label map to segmentation
            slicer.modules.segmentations.logic().ImportLabelmapToSegmentationNode(labelNode, segmentationNode)
            
            # Assign name for each segment
            segmentation = segmentationNode.GetSegmentation()
            for labelInt in uniqueLabels:
                labelName = labelNames.get(int(labelInt), f"Label_{int(labelInt)}")
                segmentId = segmentation.GetSegmentIdBySegmentName(f"Label_{int(labelInt)}")
                if segmentId:
                    segment = segmentation.GetSegment(segmentId)
                    segment.SetName(labelName)
                    
                    # Assign color
                    if int(labelInt) in colorTable:
                        r, g, b = colorTable[int(labelInt)]
                        segment.SetColor(r, g, b)
            
            # Activate 3D display
            segmentationNode.CreateClosedSurfaceRepresentation()
            displayNode = segmentationNode.GetDisplayNode()
            if displayNode:
                displayNode.SetVisibility3D(True)
                displayNode.SetVisibility2DFill(True)
                displayNode.SetVisibility2DOutline(True)
        
        # Remove original volume node
        slicer.mrmlScene.RemoveNode(volumeNode)
        
        return results
    
    def getLabelNames(self, category):
        """Returns volBrain label names - according to README.pdf."""
        labels = {}
        
        if category == "structures":
            # native_structures labels from README.pdf
            labels = {
                4: "3rd_Ventricle", 11: "4th_Ventricle",
                23: "Right_Accumbens", 30: "Left_Accumbens",
                31: "Right_Amygdala", 32: "Left_Amygdala",
                35: "Brainstem",
                36: "Right_Caudate", 37: "Left_Caudate",
                38: "Right_Cerebellum_Exterior", 39: "Left_Cerebellum_Exterior",
                40: "Right_Cerebellum_White_Matter", 41: "Left_Cerebellum_White_Matter",
                44: "Right_Cerebral_White_Matter", 45: "Left_Cerebral_White_Matter",
                47: "Right_Hippocampus", 48: "Left_Hippocampus",
                49: "Right_Inf_Lat_Vent", 50: "Left_Inf_Lat_Vent",
                51: "Right_Lateral_Ventricle", 52: "Left_Lateral_Ventricle",
                55: "Right_Pallidum", 56: "Left_Pallidum",
                57: "Right_Putamen", 58: "Left_Putamen",
                59: "Right_Thalamus", 60: "Left_Thalamus",
                61: "Right_Ventral_DC", 62: "Left_Ventral_DC",
                71: "Lobules_I-V", 72: "Lobules_VI-VII", 73: "Lobules_VIII-X",
                75: "Left_Basal_Forebrain", 76: "Right_Basal_Forebrain",
                # Cortical structures (100-207)
                100: "R_anterior_cingulate", 101: "L_anterior_cingulate",
                102: "R_anterior_insula", 103: "L_anterior_insula",
                104: "R_anterior_orbital", 105: "L_anterior_orbital",
                106: "R_angular_gyrus", 107: "L_angular_gyrus",
                108: "R_calcarine_cortex", 109: "L_calcarine_cortex",
                112: "R_central_operculum", 113: "L_central_operculum",
                114: "R_cuneus", 115: "L_cuneus",
                116: "R_entorhinal", 117: "L_entorhinal",
                118: "R_frontal_operculum", 119: "L_frontal_operculum",
                120: "R_frontal_pole", 121: "L_frontal_pole",
                122: "R_fusiform_gyrus", 123: "L_fusiform_gyrus",
                124: "R_gyrus_rectus", 125: "L_gyrus_rectus",
                128: "R_inf_occipital", 129: "L_inf_occipital",
                132: "R_inf_temporal", 133: "L_inf_temporal",
                134: "R_lingual_gyrus", 135: "L_lingual_gyrus",
                136: "R_lateral_orbital", 137: "L_lateral_orbital",
                138: "R_middle_cingulate", 139: "L_middle_cingulate",
                140: "R_medial_frontal", 141: "L_medial_frontal",
                142: "R_middle_frontal", 143: "L_middle_frontal",
                144: "R_middle_occipital", 145: "L_middle_occipital",
                146: "R_medial_orbital", 147: "L_medial_orbital",
                148: "R_postcentral_medial", 149: "L_postcentral_medial",
                150: "R_precentral_medial", 151: "L_precentral_medial",
                152: "R_sup_frontal_medial", 153: "L_sup_frontal_medial",
                154: "R_middle_temporal", 155: "L_middle_temporal",
                156: "R_occipital_pole", 157: "L_occipital_pole",
                160: "R_occipital_fusiform", 161: "L_occipital_fusiform",
                162: "R_opercular_inf_frontal", 163: "L_opercular_inf_frontal",
                164: "R_orbital_inf_frontal", 165: "L_orbital_inf_frontal",
                166: "R_posterior_cingulate", 167: "L_posterior_cingulate",
                168: "R_precuneus", 169: "L_precuneus",
                170: "R_parahippocampal", 171: "L_parahippocampal",
                172: "R_posterior_insula", 173: "L_posterior_insula",
                174: "R_parietal_operculum", 175: "L_parietal_operculum",
                176: "R_postcentral_gyrus", 177: "L_postcentral_gyrus",
                178: "R_posterior_orbital", 179: "L_posterior_orbital",
                180: "R_planum_polare", 181: "L_planum_polare",
                182: "R_precentral_gyrus", 183: "L_precentral_gyrus",
                184: "R_planum_temporale", 185: "L_planum_temporale",
                186: "R_subcallosal", 187: "L_subcallosal",
                190: "R_sup_frontal", 191: "L_sup_frontal",
                192: "R_supplementary_motor", 193: "L_supplementary_motor",
                194: "R_supramarginal", 195: "L_supramarginal",
                196: "R_sup_occipital", 197: "L_sup_occipital",
                198: "R_sup_parietal_lobule", 199: "L_sup_parietal_lobule",
                200: "R_sup_temporal", 201: "L_sup_temporal",
                202: "R_temporal_pole", 203: "L_temporal_pole",
                204: "R_triangular_inf_frontal", 205: "L_triangular_inf_frontal",
                206: "R_transverse_temporal", 207: "L_transverse_temporal"
            }
        elif category == "tissues":
            # native_tissues labels from README.pdf
            labels = {
                1: "CSF",
                2: "Cortical_GM",
                3: "Cerebrum_WM",
                4: "Subcortical_GM",
                5: "Cerebellum_GM",
                6: "Cerebellum_WM",
                7: "Brainstem"
            }
        elif category == "lobes":
            # native_lobes labels from README.pdf
            labels = {
                1: "Right_Frontal_Lobe",
                2: "Left_Frontal_Lobe",
                3: "Right_Temporal_Lobe",
                4: "Left_Temporal_Lobe",
                5: "Right_Parietal_Lobe",
                6: "Left_Parietal_Lobe",
                7: "Right_Occipital_Lobe",
                8: "Left_Occipital_Lobe",
                9: "Right_Limbic_Lobe",
                10: "Left_Limbic_Lobe",
                11: "Right_Insular_Lobe",
                12: "Left_Insular_Lobe"
            }
        elif category == "macro":
            # native_macrostructures labels from README.pdf
            labels = {
                1: "Left_Cerebrum",
                2: "Right_Cerebrum",
                3: "Left_Cerebellum",
                4: "Right_Cerebellum",
                5: "Vermal",
                6: "Brainstem"
            }
        
        return labels
    
    def getColorTable(self, category):
        """Color table for each category - according to README.pdf."""
        colors = {}
        
        if category == "structures":
            # Ventricles - Blue tones
            colors[4] = (0.2, 0.4, 0.9)   # 3rd Ventricle
            colors[11] = (0.3, 0.5, 0.95)  # 4th Ventricle
            colors[49] = (0.25, 0.45, 0.85) # Right Inf Lat Vent
            colors[50] = (0.35, 0.55, 0.9)  # Left Inf Lat Vent
            colors[51] = (0.2, 0.5, 1.0)    # Right Lateral Ventricle
            colors[52] = (0.3, 0.6, 1.0)    # Left Lateral Ventricle
            
            # Accumbens - Pink
            colors[23] = (0.9, 0.3, 0.5)
            colors[30] = (0.95, 0.35, 0.55)
            
            # Amygdala - Red
            colors[31] = (0.8, 0.2, 0.2)
            colors[32] = (0.85, 0.25, 0.25)
            
            # Brainstem - Gray
            colors[35] = (0.5, 0.5, 0.5)
            
            # Caudate - Light green
            colors[36] = (0.3, 0.7, 0.4)
            colors[37] = (0.35, 0.75, 0.45)
            
            # Cerebellum Exterior - Orange
            colors[38] = (0.9, 0.6, 0.3)
            colors[39] = (0.95, 0.65, 0.35)
            
            # Cerebellum WM - Brown
            colors[40] = (0.7, 0.5, 0.3)
            colors[41] = (0.75, 0.55, 0.35)
            
            # Cerebral WM - White/Light gray
            colors[44] = (0.9, 0.9, 0.9)
            colors[45] = (0.85, 0.85, 0.85)
            
            # Hippocampus - Yellow
            colors[47] = (0.9, 0.8, 0.2)
            colors[48] = (0.95, 0.85, 0.25)
            
            # Pallidum - Dark green
            colors[55] = (0.4, 0.6, 0.3)
            colors[56] = (0.45, 0.65, 0.35)
            
            # Putamen - Green
            colors[57] = (0.3, 0.8, 0.4)
            colors[58] = (0.35, 0.85, 0.45)
            
            # Thalamus - Purple
            colors[59] = (0.6, 0.2, 0.6)
            colors[60] = (0.65, 0.25, 0.65)
            
            # Ventral DC - Light purple
            colors[61] = (0.7, 0.4, 0.7)
            colors[62] = (0.75, 0.45, 0.75)
            
            # Cerebellum Lobules - Orange tones
            colors[71] = (0.85, 0.5, 0.2)
            colors[72] = (0.9, 0.55, 0.25)
            colors[73] = (0.95, 0.6, 0.3)
            
            # Basal Forebrain - Light pink
            colors[75] = (0.8, 0.5, 0.6)
            colors[76] = (0.85, 0.55, 0.65)
            
            # Cortical structures (100-207) - Spectrum colors
            # Frontal - Red tones
            for label in [100, 101, 104, 105, 118, 119, 120, 121, 124, 125, 136, 137, 
                         140, 141, 142, 143, 146, 147, 150, 151, 152, 153, 162, 163, 
                         164, 165, 178, 179, 182, 183, 186, 187, 190, 191, 192, 193, 
                         204, 205]:
                if label % 2 == 0:
                    colors[label] = (0.8, 0.2, 0.2)
                else:
                    colors[label] = (0.85, 0.25, 0.25)
            
            # Temporal - Blue tones
            for label in [122, 123, 132, 133, 154, 155, 180, 181, 184, 185, 200, 201, 
                         202, 203, 206, 207]:
                if label % 2 == 0:
                    colors[label] = (0.2, 0.2, 0.8)
                else:
                    colors[label] = (0.25, 0.25, 0.85)
            
            # Parietal - Green tones
            for label in [106, 107, 148, 149, 168, 169, 174, 175, 176, 177, 194, 195, 
                         198, 199]:
                if label % 2 == 0:
                    colors[label] = (0.2, 0.7, 0.3)
                else:
                    colors[label] = (0.25, 0.75, 0.35)
            
            # Occipital - Yellow tones
            for label in [108, 109, 114, 115, 128, 129, 134, 135, 144, 145, 156, 157, 
                         160, 161, 196, 197]:
                if label % 2 == 0:
                    colors[label] = (0.9, 0.8, 0.2)
                else:
                    colors[label] = (0.95, 0.85, 0.25)
            
            # Cingulate - Purple tones
            for label in [138, 139, 166, 167]:
                if label % 2 == 0:
                    colors[label] = (0.6, 0.2, 0.6)
                else:
                    colors[label] = (0.65, 0.25, 0.65)
            
            # Insula - Orange
            for label in [102, 103, 172, 173]:
                if label % 2 == 0:
                    colors[label] = (0.9, 0.5, 0.2)
                else:
                    colors[label] = (0.95, 0.55, 0.25)
            
            # Parahippocampal/Entorhinal - Light yellow
            for label in [116, 117, 170, 171]:
                if label % 2 == 0:
                    colors[label] = (0.8, 0.7, 0.3)
                else:
                    colors[label] = (0.85, 0.75, 0.35)
            
            # Central operculum - Pink
            for label in [112, 113]:
                if label % 2 == 0:
                    colors[label] = (0.9, 0.4, 0.5)
                else:
                    colors[label] = (0.95, 0.45, 0.55)
                    
        elif category == "tissues":
            colors[1] = (0.3, 0.6, 0.9)   # CSF - Blue
            colors[2] = (0.7, 0.7, 0.7)   # Cortical GM - Gray
            colors[3] = (0.9, 0.9, 0.9)   # Cerebrum WM - White
            colors[4] = (0.5, 0.7, 0.4)   # Subcortical GM - Green
            colors[5] = (0.9, 0.6, 0.3)   # Cerebellum GM - Orange
            colors[6] = (0.8, 0.5, 0.2)   # Cerebellum WM - Brown
            colors[7] = (0.5, 0.5, 0.5)   # Brainstem - Gray
            
        elif category == "lobes":
            colors[1] = (0.9, 0.3, 0.3)   # Right Frontal - Red
            colors[2] = (0.95, 0.35, 0.35) # Left Frontal
            colors[3] = (0.3, 0.3, 0.9)   # Right Temporal - Blue
            colors[4] = (0.35, 0.35, 0.95) # Left Temporal
            colors[5] = (0.3, 0.8, 0.3)   # Right Parietal - Green
            colors[6] = (0.35, 0.85, 0.35) # Left Parietal
            colors[7] = (0.9, 0.9, 0.3)   # Right Occipital - Yellow
            colors[8] = (0.95, 0.95, 0.35) # Left Occipital
            colors[9] = (0.7, 0.3, 0.7)   # Right Limbic - Purple
            colors[10] = (0.75, 0.35, 0.75) # Left Limbic
            colors[11] = (0.9, 0.5, 0.3)  # Right Insular - Orange
            colors[12] = (0.95, 0.55, 0.35) # Left Insular
            
        elif category == "macro":
            colors[1] = (0.8, 0.7, 0.7)   # Left Cerebrum - Light gray
            colors[2] = (0.75, 0.65, 0.65) # Right Cerebrum
            colors[3] = (0.9, 0.6, 0.3)   # Left Cerebellum - Orange
            colors[4] = (0.85, 0.55, 0.25) # Right Cerebellum
            colors[5] = (0.95, 0.65, 0.35) # Vermal - Light orange
            colors[6] = (0.5, 0.5, 0.5)   # Brainstem - Gray
        
        return colors

class VolBrainVolumeCalculatorTest(ScriptedLoadableModuleTest):
    """Test class."""
    
    def setUp(self):
        slicer.mrmlScene.Clear()
    
    def runTest(self):
        self.setUp()
        self.test_VolBrainVolumeCalculator1()
    
    def test_VolBrainVolumeCalculator1(self):
        self.delayDisplay("Starting test")
        self.delayDisplay('Test completed')